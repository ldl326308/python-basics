from openpyxl import Workbook
from openpyxl.utils import get_column_letter, column_index_from_string

# 创建一个新的工作表
wb = Workbook()
# 激活 worksheet
ws = wb.active

# 修改Sheet名称
ws.title = 'MySheetName'

# 根据列的数字返回字母
# print(get_column_letter(2))  # B
# 根据字母返回列的数字
# print(column_index_from_string('D'))  # 4

# 写入内容
for i in range(1, 9):  # 循环列
    for j in range(1, 9):  # 循环行
        # ws[get_column_letter(i) + j] = i * j
        print('' + get_column_letter(i) + j)

wb.save('E:\\test\\openpyxl_create.xlsx')
