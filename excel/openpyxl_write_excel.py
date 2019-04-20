# -*- coding:UTF-8 -*-
# openpyxl 写入Excel Xlsx 格式文件
from openpyxl import Workbook
from openpyxl import load_workbook
import datetime


# 创建一个新的工作表
# wb = Workbook()
# 激活 worksheet
# ws = wb.active

# 查看默认Sheet名称
# print(wb.get_sheet_names())

# 打开已有Excel
wb = load_workbook('E:\\test\\test_openpyxl.xlsx')

# 激活
ws = wb.active

# 写入数据
ws['A1'] = '听'
ws.append([1, 2, 3, 4, 5, 6])
ws['A3'] = datetime.datetime.now().strftime('%Y-%m-%d')

# 创建Sheet页
ws = wb.create_sheet('MyCreateSheet')

# 创建Sheet页，插入最开始位置
ws = wb.create_sheet('MySheet', 0)

# 选择Sheet页，写入内容
ws = wb['MySheet']
ws['A1'] = 'MySheet'

# 选择Sheet页，写入内容
ws = wb.get_sheet_by_name('MyCreateSheet')
ws['B5'] = 'MyCreateSheet'

# 输出所有Sheet名称
print(wb.sheetnames)

# 访问单个单元格
print(ws['B5'])
print(ws.cell(2, 2))

# 写入数据
ws.append(list(range(100)))

# 访问多个单元格
print('通过切片', ws['A1':'C4'])  # 通过切片
print('通过列', ws['C'])  # 通过列
print('C:D ', ws['C:D'])

# 循环所有行
print('循环所有行')
for row in ws.rows:
    for cell in row:
        print(cell.value)

wb.save('E:\\test\\test_openpyxl.xlsx')
